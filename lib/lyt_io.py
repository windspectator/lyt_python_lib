from typing import List, Callable, Union

def get_temp_root() -> str:
    """
    this function will create folder for you
    """
    import lyt_sys
    import lyt_utils
    dir = f"/tmp/{lyt_sys.get_pid()}/"
    lyt_utils.mkdir(dir)
    return dir

def _timestamp() -> str:
    import time
    return str(int(10000000000*time.time()))

def get_temp_path(name: str = None, f_depth: int = 1) -> str:
    """
    this function will create folder for you
    """
    import lyt_utils
    f_name = _timestamp() + "__" + lyt_utils.get_caller_name(depth=f_depth)
    if name is not None:
        f_name += f"__{name}"
    return get_temp_root() + f_name

def get_temp_dir_path():
    import lyt_utils
    p = get_temp_path() + "/"
    lyt_utils.mkdir(p)
    return p

class temp_file:
    def __init__(self, data: Union[List[str], str], name: str = None):
        self.f_path = get_temp_path(name, f_depth=2)
        save_txt(self.f_path, data)

    def __enter__(self):
        return self.f_path

    def __exit__(self, *_):
        import lyt_utils
        lyt_utils.rm(self.f_path, quiet=True)

class temp_file_remote:
    def __init__(self, remote_path: str, name: str = None):
        import lyt_utils
        self.f_path = get_temp_path(name, f_depth=2)
        lyt_utils.scp(remote_path, self.f_path, return_all=True)
    
    def clean(self):
        import lyt_utils
        lyt_utils.rm(self.f_path, quiet=True)

    def __enter__(self):
        return self.f_path

    def __exit__(self, *_):
        self.clean()

# same as save/load in matlab
def save(name, data):
    import pickle
    pickle.dump(data, open(name, "wb"))

def load(name, encoding="ASCII"):
    import pickle
    return pickle.load(open(name, "rb"), encoding=encoding)

def __load_txt(file, start_from=None, stop_at=None):
    lines = []
    started = True if start_from is None else False
    for line in file:
        if not started:
            if start_from in line:
                started = True
            else:
                continue
        if stop_at is not None and stop_at in line:
            break
        lines.append(line)
    return lines

def _load_txt(
        name: str, strip: bool = True, start_from: str = None, stop_at: str = None
    ) -> List[str]:
    with open(name, encoding="utf-8") as f:
        if start_from is None and stop_at is None:
            lines = f.readlines()
        else:
            lines = __load_txt(f, start_from, stop_at)
    if strip:
        lines = [x.strip() for x in lines]
    else:
        lines = [x[:-1] if x[-1] == "\n" else x for x in lines]
    return lines

# load/save strings line by line in text file.
def load_txt(
        name: str, remote: bool = False,
        strip: bool = True, start_from: str = None, stop_at: str = None
    ) -> List[str]:
    if not remote:
        return _load_txt(name, strip=strip, start_from=start_from, stop_at=stop_at)

    with temp_file_remote(name) as fname:
        return _load_txt(fname, strip=strip, start_from=start_from, stop_at=stop_at)

def load_txt_remote(
        path: str, strip: bool = True, start_from: str = None, stop_at: str = None
    ) -> List[str]:
    return load_txt(
        path, remote=True, strip=strip, start_from=start_from, stop_at=stop_at
    )

def _save_txt(name: str, data: Union[List[str], str]) -> None:
    if type(data) is str:
        data = [data]

    with open(name, 'w', encoding="utf-8") as f:
        for d in data:
            f.write(d)
            f.write('\n')

def save_txt(name: str, data: Union[List[str], str], remote: bool = False) -> None:
    if not remote:
        _save_txt(name, data=data)
        return
    
    import lyt_utils
    with temp_file(data) as fname:
        lyt_utils.scp(fname, name, return_all=True)

def insert_txt(name: str, line_index: int, data: Union[List[str], str]) -> None:
    if type(data) is str:
        data = [data]

    lines = load_txt(name, strip=False)
    lines[line_index:line_index] = data
    save_txt(name, lines)

def load_json(name):
    import json
    with open(name, 'r', encoding="utf-8") as f:
        return json.load(f)

def save_json(name, data):
    import json
    with open(name, 'w', encoding="utf-8") as f:
        json.dump(data, f, indent=4, sort_keys=True)

def load_csv(name, delimiter=","):
    data = load_txt(name)
    return [x.split(delimiter) for x in data]

def save_excel(
    path: str, data: List[List], sheet_name: str = "main", queit: bool = False
) -> None:
    """
    path should be end with .xlsx

    if you need to save only one sheet, sheet_content should be like:
        [
            [data_1_1, data_1_2, ..., data_1_n],
            [data_2_1, data_2_2, ..., data_2_n],
            ...
            [data_m_1, data_m_2, ..., data_m_n],
        ]
    Every element should be string or None.
    In this mode, you can set your single sheet's name in param sheet_name.

    If you need to save multiple sheets, use following format:
        [
            [sheet_name_1, [sheet_content_1]],
            [sheet_name_2, [sheet_content_2]],
            ...,
            [sheet_name_n, [sheet_content_n]],
        ]
    """
    import openpyxl
    import lyt_utils

    stump = lyt_utils.get_element(data, 0, 1)
    if type(stump) is str or not lyt_utils.is_iterable(stump):
        data = [[sheet_name, data]]

    tqdm = (lambda x : x) if queit else lyt_utils.tqdm
    wb = openpyxl.Workbook()
    for i, (name, value) in enumerate(data):
        if i == 0:
            ws = wb.worksheets[0]
            ws.title = name
        else:
            ws = wb.create_sheet(name)

        for j, line in tqdm(enumerate(value)):
            for k, e in enumerate(line):
                if e is None:
                    continue
                ws.cell(j + 1, k + 1).value = e

    wb.save(filename=path)

def format_folder(path):
    if not path.endswith("/"):
        path = path + "/"
    return path

def is_dir(path):
    from pathlib import Path
    return Path(path).resolve().is_dir()

def get_path(path):
    from pathlib import Path
    return Path(path).resolve().as_posix()

def get_path_name(path):
    from pathlib import Path
    return Path(path).name

def get_path_stem(path):
    from pathlib import Path
    return Path(path).stem

def get_path_parent(path, repeat=1):
    from pathlib import Path
    path = Path(path).resolve()
    for _ in range(repeat):
        path = path.parent
    return format_folder(path.as_posix())

def get_path_children(
        path: str,
        only_file: bool = False, only_dir: bool = False, recursive: bool = False,
        only_name: bool = False, pattern: str = None,
        filter_func: Callable[[str], bool] = (lambda _ : True)
    ) -> List[str]:
    """
    Do not set recursive when you use a pattern,
    eg. you should pass a pattern like "a/* /c/*.cpp" or "** /*.py" (ignore space)

    @only_name: return name but not full path
    """
    from pathlib import Path
    path = Path(path).resolve()
    result = []

    if only_name:
        def _path2str(path):
            return path.name
    else:
        def _path2str(path):
            return path.as_posix()
    if pattern is None:
        def _path_iterate(path):
            return path.iterdir()
    else:
        def _path_iterate(path):
            return path.glob(pattern)
        assert recursive is False

    if not recursive:
        for p in _path_iterate(path):
            if not filter_func(p):
                continue
            if only_file and p.is_dir():
                continue
            if only_dir and not p.is_dir():
                continue
            result.append(_path2str(p))
        return result

    def _get_path_children_recursive(result, cur_path):
        for p in cur_path.iterdir():
            if not p.is_dir():
                if not filter_func(p):
                    continue
                if not only_dir:
                    result.append(_path2str(p))
                continue

            # p is dir
            if not only_file:
                result.append(_path2str(p))
            _get_path_children_recursive(result, p)
        
        return result
    
    return _get_path_children_recursive(result, path)

def get_path_subdirs(path, **kwargs):
    return get_path_children(path, only_dir=True, **kwargs)

def get_path_subfiles(path, **kwargs):
    return get_path_children(path, only_file=True, **kwargs)

def is_path_exist_remote(path: str):
    import lyt_utils
    ip, remote_path = path.split(":", maxsplit=1)
    try:
        lyt_utils.run_remote(
            ip, f"test -e {remote_path}", print_command=False, return_all=True
        )
        return True
    except lyt_utils.Return_nonzero_exception:
        return False

def is_path_exist(path: str, remote: bool = False):
    if remote:
        return is_path_exist_remote(path)

    from pathlib import Path
    return Path(path).exists()

def assert_path_exist(path):
    if not is_path_exist(path):
        raise FileNotFoundError()

def mkdirs(path, exist_ok=True):
    """
    deprecated!
    """
    import os
    os.makedirs(path, exist_ok=exist_ok)

def edit_file_by_line(src: str, func: Callable, dst: str = None, remote: bool = False) -> int:
    if dst is None:
        dst = src

    src_lines = load_txt(src, strip=False, remote=remote)
    dst_lines = []
    for line in src_lines:
        new_line = func(line)
        if new_line is None:
            continue
        if type(new_line) is str:
            dst_lines.append(new_line)
        else:
            dst_lines.extend(new_line)
    save_txt(dst, dst_lines, remote=remote)
    return len(dst_lines) - len(src_lines)

def edit_file_by_line_remote(vm_ip, src, func, dst=None):
    if dst is None:
        dst = src
    return edit_file_by_line(f"{vm_ip}:{src}", func=func, dst=f"{vm_ip}:{dst}", remote=True)
